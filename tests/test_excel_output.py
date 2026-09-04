"""Tests for minimal Excel output helpers and the 12-sheet workbook writer."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

from config import (
    INTERACTION_ALPHAFOLD_DEFAULT,
    INTERACTION_EVIDENCE_DETAIL_DEFAULT,
    INTERACTION_NEIGHBORHOOD_DEFAULT,
    INTERACTION_SCORING_WEIGHTS_DEFAULT,
    InteractionScoringConfig,
    WordReportConfig,
)
from core.exceptions import ExcelOutputError
from core.models import BlastHit, CandidateScore, DomainHit, ProteinRecord
from analysis.interaction_scoring import (
    INTERACTION_EVIDENCE_DETAIL_V2_COLUMNS,
    InteractionScoringResult,
)
from output.report_v2 import bookmark_name
from output.excel import (
    EXCEL_COLUMNS,
    INDEX_ROWS_V2,
    records_to_dataframe,
    write_classification_workbook,
    write_records_to_excel,
)


def make_hit(
    subject_id: str,
    bitscore: float,
    evalue: float,
    source: str = "positive",
) -> BlastHit:
    """Create a BLAST hit for Excel output tests."""
    return BlastHit(
        query_id="protein_1",
        subject_id=subject_id,
        percent_identity=80.0,
        alignment_length=100,
        evalue=evalue,
        bitscore=bitscore,
        source=source,
    )


def make_record() -> ProteinRecord:
    """Create a protein record with enough fields for Excel tests."""
    record = ProteinRecord(
        protein_id="protein_1",
        description="candidate protein",
        old_locus_tag="MA_1234",
        sequence="MSTNPKPQR",
        positive_hits=[
            make_hit("positive_low", 20.0, 1e-50),
            make_hit("positive_best", 50.0, 1e-5),
            make_hit("positive_tie_best", 50.0, 1e-20),
        ],
        negative_hits=[make_hit("negative_best", 10.0, 1e-3, source="negative")],
        domains=[
            DomainHit(
                source="CDD",
                accession="cd12345",
                name="Thioredoxin_like",
                description="redox domain",
            ),
            DomainHit(
                source="Pfam",
                accession="PF00001",
                name="Domain",
                description="",
            ),
            DomainHit(
                source="CDD",
                accession="cd67890",
                name="Second_domain",
                description="second domain",
            ),
        ],
        motifs=["CXXC", "HXXH"],
        uniprot_accession="P12345",
        alphafold_url="https://example.test/model",
        notes=["reviewed", "export ready"],
    )
    score = CandidateScore(protein_id="protein_1")
    score.add_component("positive_hit", 5.0, "Positive BLAST hit found")
    score.add_component("domain_hit", 4.0, "Domain annotation found")
    record.score = score

    return record


def test_records_to_dataframe_column_order() -> None:
    """DataFrame columns should match the required Excel column order."""
    dataframe = records_to_dataframe({"protein_1": make_record()})

    assert list(dataframe.columns) == list(EXCEL_COLUMNS)
    assert list(dataframe.columns)[2] == "old_locus_tag"
    domain_count_index = list(dataframe.columns).index("domain_count")
    assert list(dataframe.columns)[domain_count_index + 1 : domain_count_index + 4] == [
        "unique_domain_count",
        "unique_domain_accessions",
        "unique_domain_names",
    ]


def test_empty_records_returns_expected_columns() -> None:
    """Empty records should return an empty DataFrame with stable columns."""
    dataframe = records_to_dataframe({})

    assert dataframe.empty
    assert list(dataframe.columns) == list(EXCEL_COLUMNS)


def test_records_to_dataframe_includes_score_fields() -> None:
    """Score details should be included in stable Excel columns."""
    dataframe = records_to_dataframe({"protein_1": make_record()})
    row = dataframe.iloc[0]

    assert row["total_score"] == 9.0
    assert row["score_components"] == "positive_hit=5.0; domain_hit=4.0"
    assert row["score_reasons"] == "Positive BLAST hit found; Domain annotation found"


def test_records_to_dataframe_without_score_uses_empty_score_fields() -> None:
    """Records without scores should use zero and blank score fields."""
    record = ProteinRecord(protein_id="protein_1")

    dataframe = records_to_dataframe({"protein_1": record})
    row = dataframe.iloc[0]

    assert row["total_score"] == 0
    assert row["score_components"] == ""
    assert row["score_reasons"] == ""
    assert row["old_locus_tag"] == ""


def test_records_to_dataframe_includes_old_locus_tag() -> None:
    """old_locus_tag should be exported immediately after description."""
    dataframe = records_to_dataframe({"protein_1": make_record()})
    row = dataframe.iloc[0]

    assert row["old_locus_tag"] == "MA_1234"
    assert list(dataframe.columns)[1:3] == ["description", "old_locus_tag"]


def test_records_to_dataframe_appends_positive_source_fields() -> None:
    """Positive source summary fields should be appended after existing columns."""
    record = ProteinRecord(
        protein_id="protein_1",
        positive_source_count=2,
        positive_sources_hit=["A", "B"],
        positive_sources_missing=["C"],
    )

    dataframe = records_to_dataframe({"protein_1": record})
    row = dataframe.iloc[0]

    assert list(dataframe.columns)[-3:] == [
        "positive_source_count",
        "positive_sources_hit",
        "positive_sources_missing",
    ]
    assert row["positive_source_count"] == 2
    assert row["positive_sources_hit"] == "A; B"
    assert row["positive_sources_missing"] == "C"


def test_records_to_dataframe_includes_negative_evidence_fields() -> None:
    """Negative hit evidence fields should be exported for classification sheets."""
    record = make_record()
    record.negative_best_identity = 45.0
    record.negative_best_query_coverage = 80.0
    record.negative_best_evalue = 1e-20
    record.negative_best_source = "Negative_source"
    record.negative_hit_strength = "strong"
    record.negative_strong_hit_count = 1
    record.negative_exclusion_reason = "excluded: strong negative hit"

    dataframe = records_to_dataframe({"protein_1": record})
    row = dataframe.iloc[0]

    assert row["negative_best_identity"] == 45.0
    assert row["negative_best_query_coverage"] == 80.0
    assert row["negative_best_evalue"] == 1e-20
    assert row["negative_best_source"] == "Negative_source"
    assert row["negative_hit_strength"] == "strong"
    assert row["negative_strong_hit_count"] == 1
    assert row["negative_exclusion_reason"] == "excluded: strong negative hit"


def test_records_to_dataframe_includes_domain_fields() -> None:
    """Domain details should be joined into stable Excel columns."""
    dataframe = records_to_dataframe({"protein_1": make_record()})
    row = dataframe.iloc[0]

    assert row["domain_sources"] == "CDD; Pfam"
    assert row["domain_names"] == "Thioredoxin_like; Domain; Second_domain"
    assert row["domain_accessions"] == "cd12345; PF00001; cd67890"
    assert row["domain_descriptions"] == "redox domain; second domain"
    assert row["domain_count"] == 3
    assert row["unique_domain_count"] == 3
    assert row["unique_domain_accessions"] == "cd12345; PF00001; cd67890"
    assert row["unique_domain_names"] == "Thioredoxin_like; Domain; Second_domain"


def test_records_to_dataframe_includes_unique_domain_summary() -> None:
    """Duplicate domain accessions and names should be summarized once."""
    record = ProteinRecord(
        protein_id="protein_1",
        domains=[
            DomainHit(source="Pfam", accession="PF00001", name="ABC"),
            DomainHit(source="Pfam", accession="PF00001", name="ABC"),
            DomainHit(source="CDD", accession="cd12345", name="CDD domain"),
        ],
    )

    dataframe = records_to_dataframe({"protein_1": record})
    row = dataframe.iloc[0]

    assert row["domain_count"] == 3
    assert row["unique_domain_count"] == 2
    assert row["unique_domain_accessions"] == "PF00001; cd12345"
    assert row["unique_domain_names"] == "ABC; CDD domain"


def test_records_to_dataframe_unique_names_skip_numeric_internal_ids() -> None:
    """Unique domain names should not include numeric-only internal ids."""
    record = ProteinRecord(
        protein_id="protein_1",
        domains=[
            DomainHit(source="Pfam", accession="PF01637.24", name="000001295"),
            DomainHit(source="Pfam", accession="PF03008.20", name="ABC_transporter"),
        ],
    )

    dataframe = records_to_dataframe({"protein_1": record})
    row = dataframe.iloc[0]

    assert row["unique_domain_names"] == "ABC_transporter"


def test_records_to_dataframe_empty_domains_are_blank() -> None:
    """Records without domains should use blank domain fields and count zero."""
    record = ProteinRecord(protein_id="protein_1", description="no domains")

    dataframe = records_to_dataframe({"protein_1": record})
    row = dataframe.iloc[0]

    assert row["domain_sources"] == ""
    assert row["domain_names"] == ""
    assert row["domain_accessions"] == ""
    assert row["domain_descriptions"] == ""
    assert row["domain_count"] == 0
    assert row["unique_domain_count"] == 0
    assert row["unique_domain_accessions"] == ""
    assert row["unique_domain_names"] == ""


def test_records_to_dataframe_selects_best_positive_hit() -> None:
    """Best positive hit should use highest bitscore, then lower e-value."""
    dataframe = records_to_dataframe({"protein_1": make_record()})
    row = dataframe.iloc[0]

    assert row["best_positive_hit"] == "positive_tie_best"
    assert row["best_positive_bitscore"] == 50.0
    assert row["best_positive_evalue"] == 1e-20


def test_records_to_dataframe_joins_motifs_and_notes() -> None:
    """Motifs and notes should be joined with a semicolon separator."""
    dataframe = records_to_dataframe({"protein_1": make_record()})
    row = dataframe.iloc[0]

    assert row["motifs"] == "CXXC; HXXH"
    assert row["notes"] == "reviewed; export ready"


def test_write_records_to_excel_creates_xlsx_file(tmp_path: Path) -> None:
    """Excel writer should create a readable xlsx file."""
    output_path = tmp_path / "reports" / "candidates.xlsx"

    result = write_records_to_excel({"protein_1": make_record()}, output_path)

    assert result == output_path.resolve()
    assert result.exists()

    dataframe = pd.read_excel(result, sheet_name="Candidates")
    assert list(dataframe.columns) == list(EXCEL_COLUMNS)
    assert dataframe.loc[0, "protein_id"] == "protein_1"


def test_write_records_to_excel_applies_simple_formatting(tmp_path: Path) -> None:
    """Excel output should include basic readability formatting."""
    output_path = tmp_path / "reports" / "formatted_candidates.xlsx"

    result = write_records_to_excel({"protein_1": make_record()}, output_path)

    workbook = load_workbook(result)
    worksheet = workbook["Candidates"]

    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref is not None
    assert worksheet["A1"].font.bold is True
    assert all(cell.font.bold for cell in worksheet[1])
    assert [cell.value for cell in worksheet[1]] == list(EXCEL_COLUMNS)
    assert worksheet.column_dimensions["B"].width >= 35


def test_write_records_to_excel_raises_excel_output_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Writer failures should be wrapped in ExcelOutputError."""

    class BrokenExcelWriter:
        """Context manager that fails while opening the workbook."""

        def __init__(self, *args: object, **kwargs: object) -> None:
            pass

        def __enter__(self) -> "BrokenExcelWriter":
            raise OSError("cannot write workbook")

        def __exit__(self, *args: object) -> None:
            pass

    monkeypatch.setattr("output.excel.pd.ExcelWriter", BrokenExcelWriter)

    with pytest.raises(ExcelOutputError, match="could not write the Excel file"):
        write_records_to_excel({"protein_1": make_record()}, tmp_path / "bad.xlsx")


# ---------------------------------------------------------------------------
# write_classification_workbook: Phase 6-8 Stage 1 unified 12-sheet layout
# ---------------------------------------------------------------------------

EXPECTED_SHEET_NAMES: tuple[str, ...] = (
    "01_Index",
    "02_Final_Score",
    "03_Candidate_Overview",
    "04_Score_Breakdown",
    "05_Sequence_Evidence",
    "06_Functional_Domain_Evidence",
    "07_Evolutionary_Evidence",
    "08_Genomic_Context",
    "09_Interaction_Evidence",
    "10_Negative_Evidence",
    "11_Raw_Audit",
    "12_Reserved",
)


def blast_classification(**buckets: dict[str, ProteinRecord]) -> SimpleNamespace:
    """Build a minimal blast_classification-like object with only the given buckets set."""
    defaults: dict[str, dict[str, ProteinRecord]] = {
        "all_records": {},
        "positive_only_records": {},
        "candidates_relaxed_records": {},
        "positive_all_sources_records": {},
        "negative_unmatched_records": {},
        "no_hit_records": {},
        "negative_hit_records": {},
        "negative_strong_hit_records": {},
        "negative_medium_hit_records": {},
        "negative_weak_hit_records": {},
    }
    defaults.update(buckets)
    return SimpleNamespace(**defaults)


def app_config(*, enabled: bool = False, scoring_model: str = "legacy_additive") -> SimpleNamespace:
    """Build a minimal app config object, matching write_classification_workbook's needs."""
    return SimpleNamespace(
        interaction_scoring=InteractionScoringConfig(
            enabled=enabled,
            query_proteins=(),
            query_fasta=None,
            candidate_sources={"candidates": True},
            max_candidates_per_query=200,
            include_sequences_in_excel=False,
            scoring_weights=INTERACTION_SCORING_WEIGHTS_DEFAULT,
            alphafold=INTERACTION_ALPHAFOLD_DEFAULT,
            neighborhood=INTERACTION_NEIGHBORHOOD_DEFAULT,
            scoring_model=scoring_model,
            scoring_engine_config=None,
            functional_complementarity_ruleset=None,
            pih_evidence_bundle=None,
            evidence_detail_sheet=INTERACTION_EVIDENCE_DETAIL_DEFAULT,
        )
    )


def _minimal_pair_row(candidate_protein_id: str = "candidate_1", **extra: object) -> dict:
    """A minimal but complete consolidated pair row for 02_Final_Score/04_Score_Breakdown tests."""
    row = {
        "query_id": "query_1",
        "query_protein_id": "query_1",
        "query_old_locus_tag": "",
        "candidate_rank": 1,
        "candidate_protein_id": candidate_protein_id,
        "candidate_old_locus_tag": "",
        "candidate_source": "Candidates",
        "candidate_description": "candidate",
        "negative_hit_strength": "none",
        "interaction_priority_score": 42.0,
        "interaction_score_reasons": "candidate source: Candidates",
        "candidate_priority_score": 30.0,
        "same_gene_neighborhood_score": 0.0,
        "distance_bp": None,
        "co_occurrence_score": 0.0,
        "domain_complementarity_score": 0.0,
        "alphafold_readiness_score": 10.0,
        "pair_total_length": 20,
        "alphafold_recommended": True,
        "protein_hunter_score": 14.0,
        "protein_hunter_score_components": "no_negative_hit=5.0; domain_hit=4.0",
        "protein_hunter_score_reasons": "This protein has no negative BLAST hits.",
        "interaction_score": 40.0,
        "final_score": 42.0,
        "final_score_tier": "Tier3_Moderate",
    }
    row.update(extra)
    return row


def test_write_classification_workbook_creates_exactly_12_sheets(tmp_path: Path) -> None:
    """The workbook should have exactly the 12 Phase 6-8 Stage 1 sheets, in order."""
    output_path = tmp_path / "reports" / "classification.xlsx"

    result = write_classification_workbook(
        config=app_config(),
        blast_classification=blast_classification(),
        output_path=output_path,
    )

    workbook = load_workbook(result)
    assert workbook.sheetnames == list(EXPECTED_SHEET_NAMES)


def test_candidate_overview_consolidates_base_classification_buckets(tmp_path: Path) -> None:
    """03_Candidate_Overview should have one row per protein with a consolidated candidate_source."""
    candidate = ProteinRecord(protein_id="A_candidate", positive_hits=[make_hit("positive", 50.0, 1e-20)])
    relaxed_only = ProteinRecord(protein_id="B_relaxed_only")
    no_hit = ProteinRecord(protein_id="C_no_hit")
    output_path = tmp_path / "reports" / "overview.xlsx"

    bc = blast_classification(
        all_records={"A_candidate": candidate, "B_relaxed_only": relaxed_only, "C_no_hit": no_hit},
        positive_only_records={"A_candidate": candidate},
        candidates_relaxed_records={"A_candidate": candidate, "B_relaxed_only": relaxed_only},
        no_hit_records={"C_no_hit": no_hit},
    )

    result = write_classification_workbook(config=app_config(), blast_classification=bc, output_path=output_path)

    overview = pd.read_excel(result, sheet_name="03_Candidate_Overview", header=1)
    by_id = overview.set_index("protein_id")
    assert by_id.loc["A_candidate", "candidate_source"] == "Candidates"
    assert by_id.loc["B_relaxed_only", "candidate_source"] == "Candidates_relaxed"
    assert by_id.loc["C_no_hit", "candidate_source"] == "No_hit"


def test_final_score_falls_back_to_protein_hunter_alone_without_interaction_result(
    tmp_path: Path,
) -> None:
    """02_Final_Score should still be populated (protein_hunter_score alone) with no interaction_result."""
    candidate = ProteinRecord(protein_id="A_candidate", positive_hits=[make_hit("positive", 50.0, 1e-20)])
    candidate.score = CandidateScore(protein_id="A_candidate")
    candidate.score.add_component("positive_hit", 5.0, "hit")
    output_path = tmp_path / "reports" / "final_score_fallback.xlsx"

    bc = blast_classification(
        all_records={"A_candidate": candidate},
        positive_only_records={"A_candidate": candidate},
    )

    result = write_classification_workbook(config=app_config(), blast_classification=bc, output_path=output_path)

    final_score = pd.read_excel(result, sheet_name="02_Final_Score", header=1)
    assert len(final_score) == 1
    assert final_score.loc[0, "candidate_protein_id"] == "A_candidate"
    assert final_score.loc[0, "interaction_score"] != final_score.loc[0, "interaction_score"] or pd.isna(
        final_score.loc[0, "interaction_score"]
    )
    assert not pd.isna(final_score.loc[0, "final_score"])


def test_final_score_uses_interaction_result_when_present(tmp_path: Path) -> None:
    """02_Final_Score should reflect interaction_result's rows, ranked by final_score."""
    output_path = tmp_path / "reports" / "final_score_interaction.xlsx"
    interaction_result = InteractionScoringResult(
        query_rows=[],
        source_rows={"Interaction_Candidates": [_minimal_pair_row()]},
        neighborhood_rows=[],
        warnings=[],
    )

    result = write_classification_workbook(
        config=app_config(enabled=True),
        blast_classification=blast_classification(),
        output_path=output_path,
        interaction_result=interaction_result,
    )

    final_score = pd.read_excel(result, sheet_name="02_Final_Score", header=1)
    assert len(final_score) == 1
    assert final_score.loc[0, "candidate_protein_id"] == "candidate_1"
    assert final_score.loc[0, "final_score"] == 42.0
    assert final_score.loc[0, "candidate_rank"] == 1


def test_final_score_dedups_candidate_scored_under_multiple_buckets(tmp_path: Path) -> None:
    """A candidate scored under both Candidates and Candidates_relaxed should appear once."""
    output_path = tmp_path / "reports" / "final_score_dedup.xlsx"
    interaction_result = InteractionScoringResult(
        query_rows=[],
        source_rows={
            "Interaction_Candidates": [_minimal_pair_row(final_score=50.0, candidate_source="Candidates")],
            "Interaction_Candidates_relaxed": [
                _minimal_pair_row(final_score=40.0, candidate_source="Candidates_relaxed")
            ],
        },
        neighborhood_rows=[],
        warnings=[],
    )

    result = write_classification_workbook(
        config=app_config(enabled=True),
        blast_classification=blast_classification(),
        output_path=output_path,
        interaction_result=interaction_result,
    )

    final_score = pd.read_excel(result, sheet_name="02_Final_Score", header=1)
    assert len(final_score) == 1
    assert final_score.loc[0, "candidate_source"] == "Candidates"
    assert final_score.loc[0, "final_score"] == 50.0


def test_category_evidence_sheets_populated_for_v2_and_empty_for_legacy(tmp_path: Path) -> None:
    """05-10 should show only their own category's rows for v2, and stay empty for legacy_additive."""
    detail_rows = [
        {
            "query_id": "query_1",
            "query_protein_id": "query_1",
            "query_old_locus_tag": "",
            "candidate_protein_id": "candidate_1",
            "candidate_old_locus_tag": "",
            "candidate_source": "Candidates",
            "candidate_rank": 1,
            "category": "genomic_context",
            "component_name": "genomic_context",
            "status": "AVAILABLE",
            "raw_value": 100,
            "normalized_value": 1.0,
            "weight": 1.0,
            "category_cap": 25.0,
            "is_negative": False,
            "explanation": "near query",
        },
        {
            "query_id": "query_1",
            "query_protein_id": "query_1",
            "query_old_locus_tag": "",
            "candidate_protein_id": "candidate_1",
            "candidate_old_locus_tag": "",
            "candidate_source": "Candidates",
            "candidate_rank": 1,
            "category": "functional_annotation",
            "component_name": "co_occurrence",
            "status": "AVAILABLE",
            "raw_value": 1.0,
            "normalized_value": 1.0,
            "weight": 10.0,
            "category_cap": 20.0,
            "is_negative": False,
            "explanation": "shared source",
        },
    ]
    output_path_v2 = tmp_path / "reports" / "category_evidence_v2.xlsx"
    interaction_result_v2 = InteractionScoringResult(
        query_rows=[],
        source_rows={"Interaction_Candidates": [_minimal_pair_row()]},
        neighborhood_rows=[],
        warnings=[],
        evidence_detail_rows=detail_rows,
        evidence_detail_scoring_model="v2_evidence_based",
    )
    result_v2 = write_classification_workbook(
        config=app_config(enabled=True, scoring_model="v2_evidence_based"),
        blast_classification=blast_classification(),
        output_path=output_path_v2,
        interaction_result=interaction_result_v2,
    )
    genomic_context = pd.read_excel(result_v2, sheet_name="08_Genomic_Context", header=1)
    functional_domain = pd.read_excel(result_v2, sheet_name="06_Functional_Domain_Evidence", header=1)
    evolutionary = pd.read_excel(result_v2, sheet_name="07_Evolutionary_Evidence", header=1)
    assert len(genomic_context) == 1
    assert genomic_context.loc[0, "component_name"] == "genomic_context"
    assert len(functional_domain) == 1
    assert functional_domain.loc[0, "component_name"] == "co_occurrence"
    assert len(evolutionary) == 0

    output_path_legacy = tmp_path / "reports" / "category_evidence_legacy.xlsx"
    interaction_result_legacy = InteractionScoringResult(
        query_rows=[],
        source_rows={"Interaction_Candidates": [_minimal_pair_row()]},
        neighborhood_rows=[],
        warnings=[],
        evidence_detail_rows=[{"query_id": "query_1", "candidate_protein_id": "candidate_1"}],
        evidence_detail_scoring_model="legacy_additive",
    )
    result_legacy = write_classification_workbook(
        config=app_config(enabled=True, scoring_model="legacy_additive"),
        blast_classification=blast_classification(),
        output_path=output_path_legacy,
        interaction_result=interaction_result_legacy,
    )
    legacy_genomic_context = pd.read_excel(result_legacy, sheet_name="08_Genomic_Context", header=1)
    assert len(legacy_genomic_context) == 0
    assert list(legacy_genomic_context.columns) == list(INTERACTION_EVIDENCE_DETAIL_V2_COLUMNS)


def test_raw_audit_sheet_stacks_detail_query_and_neighborhood_blocks(tmp_path: Path) -> None:
    """11_Raw_Audit should contain the detail table plus labeled query/neighborhood blocks below it."""
    output_path = tmp_path / "reports" / "raw_audit.xlsx"
    interaction_result = InteractionScoringResult(
        query_rows=[
            {
                "query_id": "query_1",
                "input_protein_id": "query_1",
                "input_old_locus_tag": "",
                "resolved_protein_id": "query_1",
                "resolved_old_locus_tag": "",
                "sequence_length": 10,
                "resolution_status": "resolved",
                "description": "query",
                "notes": "",
            }
        ],
        source_rows={"Interaction_Candidates": [_minimal_pair_row()]},
        neighborhood_rows=[
            {
                "query_id": "query_1",
                "candidate_protein_id": "candidate_1",
                "candidate_rank_by_distance": 1,
            }
        ],
        warnings=[],
        evidence_detail_rows=[{"query_id": "query_1", "candidate_protein_id": "candidate_1"}],
        evidence_detail_scoring_model="legacy_additive",
    )

    result = write_classification_workbook(
        config=app_config(enabled=True),
        blast_classification=blast_classification(),
        output_path=output_path,
        interaction_result=interaction_result,
    )

    workbook = load_workbook(result)
    worksheet = workbook["11_Raw_Audit"]
    values = [cell.value for row in worksheet.iter_rows() for cell in row if cell.value is not None]
    text = "\n".join(str(v) for v in values)
    assert "Interaction_query: resolved query proteins" in text
    assert "Interaction_Neighborhood: candidate-candidate genomic proximity pairs" in text
    assert "resolution_status" in text  # query block header
    assert "candidate_rank_by_distance" in text  # neighborhood block header


def test_reserved_sheet_is_present_and_empty(tmp_path: Path) -> None:
    """12_Reserved should exist with a placeholder note and no data rows."""
    output_path = tmp_path / "reports" / "reserved.xlsx"

    result = write_classification_workbook(
        config=app_config(),
        blast_classification=blast_classification(),
        output_path=output_path,
    )

    workbook = load_workbook(result)
    worksheet = workbook["12_Reserved"]
    assert "Reserved for future expansion" in str(worksheet["A2"].value)


def test_index_sheet_explains_interaction_scoring_columns(tmp_path: Path) -> None:
    """Index should include short explanations for key interaction scoring columns."""
    output_path = tmp_path / "reports" / "index_scoring_explanations.xlsx"

    result = write_classification_workbook(
        config=app_config(),
        blast_classification=blast_classification(),
        output_path=output_path,
    )

    workbook = load_workbook(result)
    index_values = [
        str(cell.value)
        for row in workbook["01_Index"].iter_rows()
        for cell in row
        if cell.value is not None
    ]
    index_text = "\n".join(index_values)

    assert "interaction_priority_score" in index_text
    assert "distance_independent_score" in index_text
    assert "protein_hunter_score" in index_text
    assert "not a direct protein-protein interaction probability" in index_text
    assert "string_ppi_score" in index_text
    assert "string-db.org" in index_text
    assert "CC BY 4.0" in index_text
    assert "functional_domain_score" in index_text
    assert "final_score_negative_penalty" in index_text
    assert "word_report_link" in index_text


def test_classification_workbook_index_links_all_sheets(tmp_path: Path) -> None:
    """01_Index should be first and link every one of the other 11 sheets."""
    output_path = tmp_path / "reports" / "classification_links.xlsx"

    result = write_classification_workbook(
        config=app_config(),
        blast_classification=blast_classification(),
        output_path=output_path,
    )

    workbook = load_workbook(result)
    assert workbook.sheetnames[0] == "01_Index"
    index = workbook["01_Index"]
    expected_sheets = [row[0] for row in INDEX_ROWS_V2]
    linked_sheets = [
        index.cell(row=row_index, column=1).value for row_index in range(2, 2 + len(expected_sheets))
    ]
    assert linked_sheets == expected_sheets
    for row_index, sheet_name in enumerate(expected_sheets, start=2):
        assert index.cell(row=row_index, column=1).hyperlink.target == f"#'{sheet_name}'!A1"

    for sheet_name in expected_sheets:
        worksheet = workbook[sheet_name]
        assert worksheet["A1"].value == "Back to Index"
        assert worksheet["A1"].hyperlink.target == "#'01_Index'!A1"


def test_index_hyperlinks_point_to_existing_sheets(tmp_path: Path) -> None:
    """Every Index hyperlink should point to a sheet that actually exists."""
    output_path = tmp_path / "reports" / "interaction_links.xlsx"

    result = write_classification_workbook(
        config=app_config(),
        blast_classification=blast_classification(),
        output_path=output_path,
    )

    workbook = load_workbook(result)
    sheet_names = set(workbook.sheetnames)
    for cell in workbook["01_Index"]["A"]:
        if cell.hyperlink is None:
            continue
        target = cell.hyperlink.target
        linked_sheet = target.split("'")[1]
        assert linked_sheet in sheet_names


# ---------------------------------------------------------------------------
# Phase 6-8 Stage 2: word_report_link (Excel -> Word cross-link)
# ---------------------------------------------------------------------------


def test_word_report_link_blank_without_filename(tmp_path: Path) -> None:
    """Default (word_report_filename=None) keeps word_report_link blank, matching Stage 1's reserved column."""
    output_path = tmp_path / "reports" / "no_word_link.xlsx"
    interaction_result = InteractionScoringResult(
        query_rows=[],
        source_rows={"Interaction_Candidates": [_minimal_pair_row()]},
        neighborhood_rows=[],
        warnings=[],
    )

    result = write_classification_workbook(
        config=app_config(enabled=True),
        blast_classification=blast_classification(),
        output_path=output_path,
        interaction_result=interaction_result,
    )

    final_score = pd.read_excel(result, sheet_name="02_Final_Score", header=1)
    assert pd.isna(final_score.loc[0, "word_report_link"])


def test_word_report_link_populated_and_clickable_when_filename_given(tmp_path: Path) -> None:
    """A candidate the Word report would show gets a real, clickable hyperlink cell."""
    output_path = tmp_path / "reports" / "word_link.xlsx"
    interaction_result = InteractionScoringResult(
        query_rows=[],
        source_rows={"Interaction_Candidates": [_minimal_pair_row(candidate_protein_id="candidate_1")]},
        neighborhood_rows=[],
        warnings=[],
    )

    result = write_classification_workbook(
        config=app_config(enabled=True),
        blast_classification=blast_classification(),
        output_path=output_path,
        interaction_result=interaction_result,
        word_report_filename="ProteinHunter_report.docx",
    )

    expected_target = f"ProteinHunter_report.docx#{bookmark_name('query_1', 'candidate_1')}"
    final_score = pd.read_excel(result, sheet_name="02_Final_Score", header=1)
    assert final_score.loc[0, "word_report_link"] == expected_target

    workbook = load_workbook(result)
    worksheet = workbook["02_Final_Score"]
    column_index = list(final_score.columns).index("word_report_link") + 1
    cell = worksheet.cell(row=3, column=column_index)
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == expected_target
    assert cell.style == "Hyperlink"


def test_word_report_link_respects_max_candidates_per_query(tmp_path: Path) -> None:
    """A candidate ranked beyond max_candidates_per_query (and not Tier1/Tier2) stays unlinked."""
    output_path = tmp_path / "reports" / "word_link_topn.xlsx"
    rows = [
        _minimal_pair_row(candidate_protein_id="in_range", final_score=90.0, final_score_tier="Tier2_Strong"),
        _minimal_pair_row(candidate_protein_id="out_of_range", final_score=10.0, final_score_tier="Tier4_Weak"),
    ]
    interaction_result = InteractionScoringResult(
        query_rows=[],
        source_rows={"Interaction_Candidates": rows},
        neighborhood_rows=[],
        warnings=[],
    )
    config = app_config(enabled=True)
    config.interaction_scoring = InteractionScoringConfig(
        **{**config.interaction_scoring.__dict__, "word_report": WordReportConfig(enabled=True, max_candidates_per_query=1)}
    )

    result = write_classification_workbook(
        config=config,
        blast_classification=blast_classification(),
        output_path=output_path,
        interaction_result=interaction_result,
        word_report_filename="report.docx",
    )

    final_score = pd.read_excel(result, sheet_name="02_Final_Score", header=1).set_index("candidate_protein_id")
    assert not pd.isna(final_score.loc["in_range", "word_report_link"])
    assert pd.isna(final_score.loc["out_of_range", "word_report_link"])
