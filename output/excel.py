"""Minimal Excel output helpers for ProteinHunter candidate records."""

from __future__ import annotations

import re
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.exceptions import ExcelOutputError
from core.models import BlastHit, ProteinRecord
from analysis.interaction_scoring import (
    INTERACTION_EVIDENCE_DETAIL_V2_COLUMNS,
    INTERACTION_QUERY_COLUMNS,
    interaction_evidence_detail_columns,
    interaction_neighborhood_columns,
)
from output.report_v2 import TIER_SAFETY_NET, bookmark_name, build_workbook_sheets, select_top_candidates_per_query


EXCEL_COLUMNS: tuple[str, ...] = (
    "protein_id",
    "description",
    "old_locus_tag",
    "total_score",
    "score_components",
    "score_reasons",
    "domain_sources",
    "domain_names",
    "domain_accessions",
    "domain_descriptions",
    "domain_count",
    "unique_domain_count",
    "unique_domain_accessions",
    "unique_domain_names",
    "sequence_length",
    "positive_hit_count",
    "negative_hit_count",
    "blast_status",
    "best_positive_hit",
    "best_positive_bitscore",
    "best_positive_evalue",
    "best_negative_hit",
    "best_negative_bitscore",
    "best_negative_evalue",
    "negative_best_identity",
    "negative_best_query_coverage",
    "negative_best_evalue",
    "negative_best_source",
    "negative_hit_strength",
    "negative_strong_hit_count",
    "negative_medium_hit_count",
    "negative_weak_hit_count",
    "negative_exclusion_reason",
    "motifs",
    "uniprot_accession",
    "alphafold_url",
    "notes",
    "positive_source_count",
    "positive_sources_hit",
    "positive_sources_missing",
)


#: Phase 6-8 Stage 1 sheet names (design spec section 25.1). Sheet 01 is the
#: Index itself (not listed in INDEX_ROWS_V2, same as the old INDEX_ROWS
#: never listing "Index"); sheet 12 is deliberately reserved/empty, see
#: claude/phase678_excel_word_redesign_investigation.md.
SHEET_FINAL_SCORE = "02_Final_Score"
SHEET_CANDIDATE_OVERVIEW = "03_Candidate_Overview"
SHEET_SCORE_BREAKDOWN = "04_Score_Breakdown"
SHEET_SEQUENCE_EVIDENCE = "05_Sequence_Evidence"
SHEET_FUNCTIONAL_DOMAIN_EVIDENCE = "06_Functional_Domain_Evidence"
SHEET_EVOLUTIONARY_EVIDENCE = "07_Evolutionary_Evidence"
SHEET_GENOMIC_CONTEXT = "08_Genomic_Context"
SHEET_INTERACTION_EVIDENCE = "09_Interaction_Evidence"
SHEET_NEGATIVE_EVIDENCE = "10_Negative_Evidence"
SHEET_RAW_AUDIT = "11_Raw_Audit"
SHEET_RESERVED = "12_Reserved"

FINAL_SCORE_COLUMNS: tuple[str, ...] = (
    "candidate_rank",
    "candidate_protein_id",
    "candidate_old_locus_tag",
    "query_id",
    "query_protein_id",
    "candidate_source",
    "negative_hit_strength",
    "final_score",
    "final_score_tier",
    "protein_hunter_score",
    "interaction_score",
    "candidate_priority_score",
    "functional_domain_score",
    "evolutionary_score",
    "same_gene_neighborhood_score",
    "interaction_evidence_score",
    "final_score_negative_penalty",
    "evidence_category_count",
    "evidence_component_count",
    "available_weight_total",
    "word_report_link",
    "candidate_description",
)

CANDIDATE_OVERVIEW_COLUMNS: tuple[str, ...] = (
    "protein_id",
    "old_locus_tag",
    "description",
    "candidate_source",
    "negative_hit_strength",
    "protein_hunter_score",
    "blast_status",
    "positive_hit_count",
    "negative_hit_count",
    "best_positive_hit",
    "domain_count",
    "uniprot_accession",
    "alphafold_url",
    "positive_source_count",
    "positive_sources_hit",
    "positive_sources_missing",
    "sequence_length",
    "notes",
)

SCORE_BREAKDOWN_COLUMNS: tuple[str, ...] = (
    "query_id",
    "query_protein_id",
    "candidate_rank",
    "candidate_protein_id",
    "candidate_old_locus_tag",
    "candidate_source",
    "negative_hit_strength",
    "candidate_priority_score",
    "same_gene_neighborhood_score",
    "functional_domain_score",
    "co_occurrence_score",
    "domain_complementarity_score",
    "evolutionary_score",
    "cellular_compatibility_score",
    "interaction_evidence_score",
    "alphafold_readiness_score",
    "string_ppi_score",
    "interaction_priority_score",
    "interaction_score",
    "final_score",
    "evidence_tier",
    "interaction_evidence_tier",
    "final_score_tier",
    "formal_score_available",
    "evidence_category_count",
    "evidence_component_count",
    "available_weight_total",
    "scoring_model",
)

#: (sheet name, evidence-detail categories shown there). v2_evidence_based
#: only -- Interaction_Evidence_Detail's long format (category/component per
#: row) is the only place these categories exist as discrete rows; legacy_additive
#: has no per-category concept at all, so these sheets stay header-only for
#: that scoring model (see 11_Raw_Audit for legacy's own wide-format detail
#: instead). 07_Evolutionary_Evidence also carries pih_cellular_compatibility --
#: both are PihEvidenceBundle-sourced and, per the Stage 1 directive, both
#: are "mostly unfulfilled" today (no dedicated Cellular_Compatibility sheet
#: exists in the 12-sheet budget); cellular_compatibility_score remains
#: available as its own reference column on 04_Score_Breakdown.
CATEGORY_EVIDENCE_SHEETS: tuple[tuple[str, frozenset[str]], ...] = (
    (SHEET_SEQUENCE_EVIDENCE, frozenset({"source_classification"})),
    (SHEET_FUNCTIONAL_DOMAIN_EVIDENCE, frozenset({"functional_annotation"})),
    (SHEET_EVOLUTIONARY_EVIDENCE, frozenset({"pih_evolutionary", "pih_cellular_compatibility"})),
    (SHEET_GENOMIC_CONTEXT, frozenset({"genomic_context"})),
    (
        SHEET_INTERACTION_EVIDENCE,
        frozenset({"external_ppi_evidence", "coexpression_evidence", "pih_direct_interaction"}),
    ),
    (SHEET_NEGATIVE_EVIDENCE, frozenset({"source_reliability"})),
)


INDEX_ROWS_V2: tuple[tuple[str, str, str, str], ...] = (
    (
        SHEET_FINAL_SCORE,
        "every candidate with a resolved query (or, with interaction_scoring "
        "disabled/no query configured, every classified candidate)",
        "final-score-first ranked view: protein_hunter_score + interaction_score "
        "combined, one row per candidate_source-consolidated (query, candidate) pair",
        "start here -- primary ranked candidate list",
    ),
    (
        SHEET_CANDIDATE_OVERVIEW,
        "every classified protein",
        "global candidate identity, independent of any query: consolidated "
        "candidate_source, negative_hit_strength, and BLAST/domain summary",
        "look up a specific candidate's classification regardless of query",
    ),
    (
        SHEET_SCORE_BREAKDOWN,
        "same rows as 02_Final_Score",
        "full per-category score breakdown behind each row's final_score",
        "audit why a candidate scored the way it did",
    ),
    (
        SHEET_SEQUENCE_EVIDENCE,
        "scoring_model: v2_evidence_based only",
        "source_classification/sequence_evidence component-level detail",
        "inspect the raw BLAST-identity evidence behind the Sequence category",
    ),
    (
        SHEET_FUNCTIONAL_DOMAIN_EVIDENCE,
        "scoring_model: v2_evidence_based only",
        "functional_annotation (co_occurrence + domain_complementarity) component-level detail",
        "inspect the raw functional/domain evidence behind the Functional+Domain category",
    ),
    (
        SHEET_EVOLUTIONARY_EVIDENCE,
        "scoring_model: v2_evidence_based only, and only when a PIH evidence "
        "bundle is configured",
        "pih_evolutionary + pih_cellular_compatibility component-level detail -- "
        "mostly empty today, not a bug (see claude/phase678_excel_word_redesign_investigation.md)",
        "inspect PIH-bridged evolutionary/cellular-compatibility evidence when available",
    ),
    (
        SHEET_GENOMIC_CONTEXT,
        "scoring_model: v2_evidence_based only",
        "genomic_context (+ STRING neighborhood) component-level detail",
        "inspect the raw gene-neighborhood/STRING-neighborhood evidence",
    ),
    (
        SHEET_INTERACTION_EVIDENCE,
        "scoring_model: v2_evidence_based only",
        "external_ppi_evidence + coexpression_evidence + pih_direct_interaction "
        "component-level detail",
        "inspect the raw STRING/coexpression/PIH direct-interaction evidence",
    ),
    (
        SHEET_NEGATIVE_EVIDENCE,
        "scoring_model: v2_evidence_based only",
        "negative_hit_strength component detail plus the reserved, always "
        "NOT_APPLICABLE final_score_negative_penalty audit slot -- mostly "
        "empty today, not a bug",
        "inspect phylogenetic-novelty evidence and confirm the negative-penalty slot is unused",
    ),
    (
        SHEET_RAW_AUDIT,
        "every interaction_scoring row",
        "full unfiltered Interaction_Evidence_Detail rows, plus the resolved "
        "query proteins and candidate-candidate genomic neighborhood pairs "
        "as separate blocks further down the same sheet",
        "full audit trail / raw export",
    ),
    (
        SHEET_RESERVED,
        "n/a",
        "reserved for future expansion",
        "not used in this stage",
    ),
)


NEGATIVE_EVIDENCE_EXPLANATIONS: tuple[tuple[str, str], ...] = (
    ("negative_best_identity", "best identity among negative hits"),
    (
        "negative_best_query_coverage",
        "query coverage of the representative/best negative hit",
    ),
    ("negative_best_evalue", "E-value of the representative/best negative hit"),
    (
        "negative_best_source",
        "negative source where the representative/best negative hit was found",
    ),
    ("negative_hit_strength", "strong / medium / weak / none"),
    (
        "negative_strong_hit_count",
        "number of negative hits classified as strong",
    ),
    (
        "negative_medium_hit_count",
        "number of negative hits classified as medium",
    ),
    ("negative_weak_hit_count", "number of negative hits classified as weak"),
    (
        "negative_exclusion_reason",
        "reason why the record was retained or excluded under the current ortholog_filter mode",
    ),
)


INTERACTION_SCORE_EXPLANATIONS: tuple[tuple[str, str], ...] = (
    (
        "interaction_priority_score",
        "Overall interaction/functional priority score including candidate source, gene neighborhood, co-occurrence, domain complementarity, and modeling readiness.",
    ),
    (
        "candidate_priority_score",
        "Score based on which candidate source sheet the protein came from, such as Candidates, Candidates_relaxed, No_hit, etc.",
    ),
    (
        "same_gene_neighborhood_score",
        "Score based on genomic distance from the query protein using GFF coordinates. Close genes get positive score, but distant genes are not excluded.",
    ),
    (
        "distance_bp",
        "Genomic distance in base pairs between query and candidate genes when both coordinates are available.",
    ),
    (
        "co_occurrence_score",
        "Score based on similarity of positive/negative source hit patterns between query and candidate.",
    ),
    (
        "domain_complementarity_score",
        "Score based on meaningful functional terms from description/Pfam/CDD/domain annotations. Generic words alone are ignored.",
    ),
    (
        "alphafold_readiness_score",
        "Score indicating whether the pair is practical to model structurally based on sequence availability and length. This is not evidence of interaction.",
    ),
    (
        "distance_independent_score",
        "Score excluding gene neighborhood and AlphaFold readiness. Formula: candidate_priority_score + co_occurrence_score + domain_complementarity_score.",
    ),
    (
        "distance_independent_rank",
        "Rank based on distance_independent_score within each Interaction_* sheet.",
    ),
    (
        "priority_group",
        "Category such as nearby_candidate, distant_cooccurrence_candidate, distant_domain_candidate, no_hit_candidate, or general_candidate.",
    ),
    (
        "interaction_score_reasons",
        "Human-readable reasons explaining why the candidate received its score.",
    ),
    (
        "protein_hunter_score",
        "Query-independent 'is this generally a good candidate' score for the candidate "
        "protein alone (same formula as the Candidates sheet's total_score, but computed "
        "for every interaction_scoring candidate source, not just Candidates). Does not "
        "change if the query changes, and is never part of interaction_priority_score, "
        "candidate_rank, or sheet sort order -- reference only.",
    ),
    (
        "interaction_score",
        "Query-specific evidence only, re-normalized to 0-100: genomic_context "
        "(+ STRING's neighborhood channel) + domain_complementarity + external_ppi_evidence "
        "(STRING's cooccurrence channel) + coexpression_evidence (GSE64349 measured transcript "
        "coexpression only -- GSE77738 is computed and still shown in "
        "Interaction_Evidence_Detail, but excluded from this sum; a real-data check found it "
        "scored AlphaFold3-confirmed non-interacting pairs higher, on average, than curated true "
        "positive pairs, see claude/experimental_interactions_calibration_report.md). Populated "
        "for both scoring models -- v2_evidence_based "
        "computes it from EvidenceComponent categories (blank/None means no query-specific "
        "evidence was available at all, not a scored zero); legacy_additive computes an "
        "equivalent blended sum (always a number, 0 when there is no evidence, since legacy has "
        "no 'missing vs. evaluated-zero' concept; legacy_additive does not yet include "
        "coexpression_evidence -- v2_evidence_based only, see "
        "claude/phase6b_coexpression_design.md). Deliberately excludes source_classification, "
        "sequence_evidence, and co_occurrence, which mainly reflect the candidate's own "
        "conservation profile rather than evidence specific to this query pair. Reference only -- "
        "does not affect interaction_priority_score, candidate_rank, or sheet sort order (unless "
        "interaction_scoring.ranking_metric is set to 'interaction_score').",
    ),
    (
        "final_score",
        "Final Score (design spec sections 17-22/27): combines protein_hunter_score and "
        "interaction_score into one 0-100 value via two top-level categories -- "
        "'protein_hunter' (cap 30, protein_hunter_score normalized against its own theoretical "
        "ceiling of 18) and 'interaction' (cap 70, interaction_score/100). Both caps are "
        "PROVISIONAL, see claude/final_score_integration_investigation.md for the real-data "
        "check behind the 30/70 split. Deliberately does NOT apply a negative_hit_strength "
        "penalty here (tried and removed -- negative_hit_strength measures phylogenetic "
        "novelty, not the design spec section 7.7 'biological contradiction' concept this "
        "score's negative-evidence slot is reserved for; conflating the two collapsed the "
        "separation between true positive and AlphaFold3-negative pairs in a real-data check, "
        "since being well-conserved is exactly what routes a candidate into the Negative_hit "
        "bucket -- see claude/final_score_integration_investigation.md). "
        "interaction_priority_score's own, separate use of negative_hit_strength is unaffected. "
        "When interaction_score is unavailable for a pair, Final Score falls back "
        "to protein_hunter_score alone (re-normalized against just its own cap), the same "
        "'renormalize against whatever evidence is available' behavior every other v2 category "
        "already uses. Populated for both scoring models. Reference only by default -- does not "
        "affect interaction_priority_score, evidence_tier, candidate_rank, or sheet sort order "
        "unless interaction_scoring.ranking_metric is set to 'final_score'.",
    ),
    (
        "final_score_tier",
        "Confidence tier for final_score, using the same Tier1_VeryStrong/Tier2_Strong/"
        "Tier3_Moderate/Tier4_Weak/Unclassified thresholds as evidence_tier/"
        "interaction_evidence_tier, applied to final_score's own score/category-count instead. "
        "A separate classification from evidence_tier -- never overwrites it.",
    ),
    (
        "string_ppi_score",
        "legacy_additive only (blank for v2_evidence_based, which reports STRING evidence as "
        "separate string_cooccurrence/string_neighborhood rows in Interaction_Evidence_Detail "
        "instead). Averages STRING's cooccurrence and neighborhood channels for this pair into "
        "one 0-weights.external_ppi point score. 0 when interaction_scoring.string_ppi_ncbi_taxon_id "
        "is unset or STRING has no data for this pair -- reference only, folded into both "
        "interaction_priority_score and interaction_score.",
    ),
    (
        "candidate_source",
        "Phase 6-8 sheet redesign (design spec section 25.1): which candidate_sources bucket "
        "this candidate was classified into, consolidated to one value per (query, candidate) "
        "pair (or per candidate on 03_Candidate_Overview) -- Candidates > Positive_all_sources > "
        "Candidates_relaxed > No_hit > Negative_unmatched > Negative_hit, highest-priority bucket "
        "wins when a candidate was scored under more than one enabled bucket. "
        "Negative_strong_hit/Negative_medium_hit/Negative_weak_hit are folded into Negative_hit "
        "here -- see negative_hit_strength for the sub-classification.",
    ),
    (
        "negative_hit_strength",
        "strong / medium / weak / none -- see the Negative evidence columns block below for "
        "the full explanation. Exposed directly on every pair row (both scoring models) as of "
        "the Phase 6-8 sheet redesign, replacing the former separate Negative_strong/medium/"
        "weak_hit sheets.",
    ),
    (
        "functional_domain_score",
        "scoring_model: v2_evidence_based only. The functional_annotation category's own "
        "capped_score (co_occurrence + domain_complementarity combined and cap-renormalized) -- "
        "blank when neither component had any evidence for this pair, not a scored zero.",
    ),
    (
        "evolutionary_score",
        "scoring_model: v2_evidence_based only, and only when a PIH evidence bundle is "
        "configured. The pih_evolutionary category's own capped_score -- blank (not a scored "
        "zero) when no PIH evolutionary evidence was evaluated for this pair. Mostly blank "
        "today; not a bug, see claude/phase678_excel_word_redesign_investigation.md.",
    ),
    (
        "cellular_compatibility_score",
        "scoring_model: v2_evidence_based only, and only when a PIH evidence bundle is "
        "configured. The pih_cellular_compatibility category's own capped_score -- blank (not "
        "a scored zero) when no PIH cellular-compatibility evidence was evaluated. Mostly "
        "blank today; not a bug.",
    ),
    (
        "interaction_evidence_score",
        "scoring_model: v2_evidence_based only. Sum of the external_ppi_evidence + "
        "coexpression_evidence + pih_direct_interaction categories' own capped_score values -- "
        "a raw category-level reference (like candidate_priority_score/"
        "same_gene_neighborhood_score above), NOT the same number as interaction_score (which "
        "is re-normalized to 0-100 over a different, overlapping component scope; see "
        "interaction_score's own explanation). Blank when none of the three categories had any "
        "evidence for this pair.",
    ),
    (
        "final_score_negative_penalty",
        "Always blank/NOT_APPLICABLE. Reserved audit column for a future biological-"
        "contradiction negative-evidence signal (design spec section 7.7) -- see final_score's "
        "own explanation for why negative_hit_strength was tried here and removed. Kept as a "
        "column so a future real signal has a place to land without another sheet-layout change.",
    ),
    (
        "word_report_link",
        "Always blank in this stage (Phase 6-8 Stage 1, Excel only). Reserved for a hyperlink "
        "to this candidate's section in the single-file Word report, planned for Stage 2.",
    ),
)

INTERACTION_SCORE_NOTES: tuple[str, ...] = (
    "interaction_priority_score is not a direct protein-protein interaction probability.",
    "Gene neighborhood is used as positive evidence only.",
    "Distant archaeal candidates should not be excluded solely because they are far from the query gene.",
    "AlphaFold readiness does not mean AlphaFold predicts interaction.",
    "string_cooccurrence/string_neighborhood/string_ppi_score are derived from STRING "
    "(https://string-db.org), used under the Creative Commons Attribution 4.0 "
    "International (CC BY 4.0) license. Please credit STRING if these values are "
    "published or redistributed.",
    "coexpression_gse77738/coexpression_gse64349 are derived from NCBI GEO datasets "
    "GSE77738 (PMID 27852217) and GSE64349 (PMID 25691524), public NIH data with no "
    "reuse restriction; citing these studies when the values are published or "
    "redistributed is standard scientific courtesy, not a license requirement.",
)


def records_to_dataframe(records: dict[str, ProteinRecord]) -> pd.DataFrame:
    """Convert ProteinRecord objects into a tabular pandas DataFrame."""
    rows = [_record_to_row(record) for record in records.values()]
    return pd.DataFrame(rows, columns=EXCEL_COLUMNS)


def write_records_to_excel(
    records: dict[str, ProteinRecord],
    output_path: str | Path,
    sheet_name: str = "Candidates",
) -> Path:
    """Write ProteinRecord objects to an Excel workbook and return its path."""
    resolved_output = Path(output_path).expanduser().resolve()
    resolved_output.parent.mkdir(parents=True, exist_ok=True)
    dataframe = records_to_dataframe(records)

    try:
        with pd.ExcelWriter(resolved_output, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            _format_worksheet(worksheet, dataframe)
    except Exception as exc:
        message = (
            f"ProteinHunter could not write the Excel file: {resolved_output}. "
            "Please check that the folder is writable and the file is not open."
        )
        raise ExcelOutputError(message) from exc

    return resolved_output


def write_classification_workbook(
    config: Any,
    blast_classification: Any,
    output_path: str | Path,
    interaction_result: Any | None = None,
    word_report_filename: str | None = None,
) -> Path:
    """Write the unified 12-sheet Phase 6-8 Stage 1 workbook and return its path.

    Replaces the former ~10 base classification sheets + up to 11
    Interaction_* bucket sheets with a fixed 12-sheet layout (design spec
    section 25.1) keyed by a consolidated ``candidate_source`` column
    instead of one sheet per bucket -- see output/report_v2.py for the row
    consolidation logic and claude/phase678_excel_word_redesign_investigation.md
    for the design background.

    ``word_report_filename`` (Phase 6-8 Stage 2) is the Word report's
    filename relative to this workbook -- when given, every
    ``02_Final_Score`` row selected for the Word report (same
    ``select_top_candidates_per_query`` selection ``output/word_report.py``
    itself uses) gets a real, clickable ``word_report_link`` hyperlink into
    that candidate's bookmarked section
    (``f"{word_report_filename}#{bookmark_name(...)}"``, design spec's
    Excel-\\>Word cross-link, option A+C from
    claude/phase678_excel_word_redesign_investigation.md item 6). Left
    ``None`` (the default), every row's ``word_report_link`` stays blank,
    exactly as it has since Phase 6-8 Stage 1 reserved the column.
    """
    resolved_output = Path(output_path).expanduser().resolve()
    resolved_output.parent.mkdir(parents=True, exist_ok=True)

    sheets_data = build_workbook_sheets(config, blast_classification, interaction_result)
    scoring_model = sheets_data["evidence_detail_scoring_model"]

    if word_report_filename:
        _attach_word_report_links(sheets_data["final_score_rows"], config, word_report_filename)

    final_score_df = pd.DataFrame(sheets_data["final_score_rows"], columns=FINAL_SCORE_COLUMNS)
    overview_df = pd.DataFrame(sheets_data["overview_rows"], columns=CANDIDATE_OVERVIEW_COLUMNS)
    score_breakdown_df = pd.DataFrame(sheets_data["final_score_rows"], columns=SCORE_BREAKDOWN_COLUMNS)

    evidence_detail_rows = sheets_data["evidence_detail_rows"]
    category_dataframes: dict[str, pd.DataFrame] = {}
    for sheet_name, categories in CATEGORY_EVIDENCE_SHEETS:
        if scoring_model == "v2_evidence_based":
            filtered_rows = [row for row in evidence_detail_rows if row.get("category") in categories]
        else:
            filtered_rows = []
        category_dataframes[sheet_name] = pd.DataFrame(filtered_rows, columns=INTERACTION_EVIDENCE_DETAIL_V2_COLUMNS)

    raw_audit_df = pd.DataFrame(
        evidence_detail_rows, columns=interaction_evidence_detail_columns(scoring_model)
    )
    query_df = pd.DataFrame(sheets_data["query_rows"], columns=INTERACTION_QUERY_COLUMNS)
    neighborhood_df = pd.DataFrame(sheets_data["neighborhood_rows"], columns=interaction_neighborhood_columns())
    reserved_df = pd.DataFrame(columns=("Reserved",))

    try:
        with pd.ExcelWriter(resolved_output, engine="openpyxl") as writer:
            index_dataframe = _index_dataframe(INDEX_ROWS_V2)
            index_dataframe.to_excel(writer, sheet_name="01_Index", index=False)
            _format_index_worksheet(writer.sheets["01_Index"], index_dataframe, INDEX_ROWS_V2)

            simple_sheets: dict[str, pd.DataFrame] = {
                SHEET_FINAL_SCORE: final_score_df,
                SHEET_CANDIDATE_OVERVIEW: overview_df,
                SHEET_SCORE_BREAKDOWN: score_breakdown_df,
                **category_dataframes,
            }
            for sheet_name, dataframe in simple_sheets.items():
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                worksheet = writer.sheets[sheet_name]
                _add_back_to_index_link(worksheet)
                _format_worksheet(worksheet, dataframe, header_row=2)
                if sheet_name == SHEET_FINAL_SCORE:
                    _style_word_report_link_column(worksheet, dataframe, header_row=2)

            _write_raw_audit_sheet(writer, SHEET_RAW_AUDIT, raw_audit_df, query_df, neighborhood_df)

            reserved_df.to_excel(writer, sheet_name=SHEET_RESERVED, index=False, startrow=1)
            reserved_worksheet = writer.sheets[SHEET_RESERVED]
            _add_back_to_index_link(reserved_worksheet)
            reserved_worksheet["A2"] = "Reserved for future expansion (Word report cross-links, Stage 2)."
    except Exception as exc:
        message = (
            f"ProteinHunter could not write the Excel file: {resolved_output}. "
            "Please check that the folder is writable and the file is not open."
        )
        raise ExcelOutputError(message) from exc

    return resolved_output


def _write_raw_audit_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    detail_df: pd.DataFrame,
    query_df: pd.DataFrame,
    neighborhood_df: pd.DataFrame,
) -> None:
    """Write 11_Raw_Audit's three stacked blocks: detail, query, neighborhood."""
    detail_header_row = 2
    detail_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=detail_header_row - 1)
    worksheet = writer.sheets[sheet_name]
    _add_back_to_index_link(worksheet)
    _format_worksheet(worksheet, detail_df, header_row=detail_header_row)

    query_label_row = detail_header_row + len(detail_df) + 2
    query_header_row = query_label_row + 1
    worksheet.cell(row=query_label_row, column=1, value="Interaction_query: resolved query proteins")
    worksheet.cell(row=query_label_row, column=1).font = Font(bold=True)
    query_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=query_header_row - 1)
    for cell in worksheet[query_header_row]:
        cell.font = Font(bold=True)

    neighborhood_label_row = query_header_row + len(query_df) + 2
    neighborhood_header_row = neighborhood_label_row + 1
    worksheet.cell(
        row=neighborhood_label_row,
        column=1,
        value="Interaction_Neighborhood: candidate-candidate genomic proximity pairs",
    )
    worksheet.cell(row=neighborhood_label_row, column=1).font = Font(bold=True)
    neighborhood_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=neighborhood_header_row - 1)
    for cell in worksheet[neighborhood_header_row]:
        cell.font = Font(bold=True)


def _format_worksheet(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    header_row: int = 1,
) -> None:
    """Apply simple readability formatting to an Excel worksheet."""
    worksheet.freeze_panes = f"A{header_row + 1}"
    if len(dataframe.columns) > 0:
        last_column = get_column_letter(len(dataframe.columns))
        last_row = max(header_row, header_row + len(dataframe.index))
        worksheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"

    for header_cell in worksheet[header_row]:
        header_cell.font = Font(bold=True)

    wrap_columns = {
        "description",
        "domain_descriptions",
        "score_reasons",
        "notes",
        "positive_sources_hit",
        "positive_sources_missing",
        "negative_exclusion_reason",
        "explanation",
        "Selection rule",
        "Biological interpretation",
        "Recommended use",
    }

    for column_index, column_name in enumerate(dataframe.columns, start=1):
        column_letter = get_column_letter(column_index)
        width = _column_width(column_name, dataframe[column_name])

        if column_name in {"description", "domain_descriptions", "score_reasons"}:
            width = max(width, 35)
        elif column_name == "notes":
            width = max(width, 45)
        elif column_name in {
            "Selection rule",
            "Biological interpretation",
            "Recommended use",
        }:
            width = min(max(width, 28), 70)

        worksheet.column_dimensions[column_letter].width = width

        if column_name in wrap_columns:
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _index_dataframe(
    index_rows: tuple[tuple[str, str, str, str], ...],
) -> pd.DataFrame:
    """Return the workbook navigation index DataFrame."""
    rows: list[dict[str, str]] = [
        {
            "Sheet": sheet,
            "Selection rule": selection_rule,
            "Biological interpretation": interpretation,
            "Recommended use": recommended_use,
        }
        for sheet, selection_rule, interpretation, recommended_use in index_rows
    ]
    rows.append(
        {
            "Sheet": "",
            "Selection rule": "",
            "Biological interpretation": "",
            "Recommended use": "",
        }
    )
    rows.append(
        {
            "Sheet": "Interaction scoring columns",
            "Selection rule": "",
            "Biological interpretation": "",
            "Recommended use": "",
        }
    )
    for column_name, explanation in INTERACTION_SCORE_EXPLANATIONS:
        rows.append(
            {
                "Sheet": column_name,
                "Selection rule": explanation,
                "Biological interpretation": "",
                "Recommended use": "",
            }
        )
    rows.append(
        {
            "Sheet": "Interaction scoring notes",
            "Selection rule": "; ".join(INTERACTION_SCORE_NOTES),
            "Biological interpretation": "",
            "Recommended use": "",
        }
    )
    rows.append(
        {
            "Sheet": "",
            "Selection rule": "",
            "Biological interpretation": "",
            "Recommended use": "",
        }
    )
    rows.append(
        {
            "Sheet": "Negative evidence columns",
            "Selection rule": "",
            "Biological interpretation": "",
            "Recommended use": "",
        }
    )
    for column_name, explanation in NEGATIVE_EVIDENCE_EXPLANATIONS:
        rows.append(
            {
                "Sheet": column_name,
                "Selection rule": explanation,
                "Biological interpretation": "",
                "Recommended use": "",
            }
        )

    return pd.DataFrame(
        rows,
        columns=(
            "Sheet",
            "Selection rule",
            "Biological interpretation",
            "Recommended use",
        ),
    )


def _format_index_worksheet(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    index_rows: tuple[tuple[str, str, str, str], ...],
) -> None:
    """Apply navigation links and readable formatting to the Index sheet."""
    _format_worksheet(worksheet, dataframe)
    for row_index, sheet_name in enumerate(
        (row[0] for row in index_rows),
        start=2,
    ):
        cell = worksheet.cell(row=row_index, column=1)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.style = "Hyperlink"

    for row in worksheet.iter_rows(min_row=2):
        first_cell = row[0]
        if first_cell.value in {
            "Interaction scoring columns",
            "Interaction scoring notes",
            "Negative evidence columns",
        }:
            first_cell.font = Font(bold=True)


def _add_back_to_index_link(worksheet: Worksheet) -> None:
    """Add a consistent internal link back to the Index sheet."""
    worksheet["A1"] = "Back to Index"
    worksheet["A1"].hyperlink = "#'01_Index'!A1"
    worksheet["A1"].style = "Hyperlink"


def _attach_word_report_links(
    rows: list[dict[str, Any]], config: Any, word_report_filename: str
) -> None:
    """Populate ``word_report_link`` in place for every row the Word report also shows.

    Uses the exact same selection (``select_top_candidates_per_query``) and
    bookmark naming (``bookmark_name``) the Word report itself uses --
    both pure, deterministic functions of the row data alone (see
    output/report_v2.py), so this can recompute "would this candidate be
    in the Word report" independently, without output/excel.py and
    output/word_report.py needing to pass data to each other. Rows not
    selected keep ``word_report_link`` as whatever build_workbook_sheets
    already set it to (``None``/blank).
    """
    scoring_config = getattr(config, "interaction_scoring", None)
    word_report_config = getattr(scoring_config, "word_report", None)
    max_per_query = int(getattr(word_report_config, "max_candidates_per_query", 15))

    selected = select_top_candidates_per_query(rows, max_per_query, TIER_SAFETY_NET)
    selected_keys = {(str(row["query_id"]), str(row["candidate_protein_id"])) for row in selected}

    for row in rows:
        key = (str(row.get("query_id")), str(row.get("candidate_protein_id")))
        if key in selected_keys:
            name = bookmark_name(*key)
            row["word_report_link"] = f"{word_report_filename}#{name}"


def _style_word_report_link_column(worksheet: Worksheet, dataframe: pd.DataFrame, header_row: int) -> None:
    """Turn populated word_report_link cells into real, clickable Excel hyperlinks.

    ``dataframe.to_excel`` only ever writes plain text -- this applies the
    same hyperlink + "Hyperlink" style treatment _add_back_to_index_link
    already uses elsewhere in this file, cell by cell, to just this one
    column.
    """
    if "word_report_link" not in dataframe.columns:
        return
    column_index = list(dataframe.columns).index("word_report_link") + 1
    for offset, value in enumerate(dataframe["word_report_link"]):
        if not value or pd.isna(value):
            continue
        cell = worksheet.cell(row=header_row + 1 + offset, column=column_index)
        cell.hyperlink = str(value)
        cell.style = "Hyperlink"


def _column_width(column_name: str, values: pd.Series) -> int:
    """Return a practical Excel column width based on header and cell text."""
    max_length = len(column_name)

    for value in values:
        if pd.isna(value):
            continue

        max_length = max(max_length, len(str(value)))

    return min(max(max_length + 2, 10), 60)


def _record_to_row(record: ProteinRecord) -> dict[str, Any]:
    """Convert one ProteinRecord into one Excel row dictionary."""
    best_positive = _best_hit(record.positive_hits)
    best_negative = _best_hit(record.negative_hits)

    return {
        "protein_id": record.protein_id,
        "description": record.description,
        "old_locus_tag": record.old_locus_tag or "",
        "total_score": record.score.total_score if record.score else 0,
        "score_components": _score_components(record),
        "score_reasons": "; ".join(record.score.reasons) if record.score else "",
        "domain_sources": _unique_join(domain.source for domain in record.domains),
        "domain_names": _join(domain.name for domain in record.domains),
        "domain_accessions": _join(domain.accession for domain in record.domains),
        "domain_descriptions": _join(
            domain.description for domain in record.domains if domain.description
        ),
        "domain_count": len(record.domains),
        "unique_domain_count": _unique_domain_count(record),
        "unique_domain_accessions": _unique_join(
            domain.accession for domain in record.domains
        ),
        "unique_domain_names": _unique_domain_names(record),
        "sequence_length": record.length,
        "positive_hit_count": len(record.positive_hits),
        "negative_hit_count": len(record.negative_hits),
        "blast_status": _blast_status(record),
        "best_positive_hit": best_positive.subject_id if best_positive else None,
        "best_positive_bitscore": best_positive.bitscore if best_positive else None,
        "best_positive_evalue": best_positive.evalue if best_positive else None,
        "best_negative_hit": best_negative.subject_id if best_negative else None,
        "best_negative_bitscore": best_negative.bitscore if best_negative else None,
        "best_negative_evalue": best_negative.evalue if best_negative else None,
        "negative_best_identity": record.negative_best_identity,
        "negative_best_query_coverage": record.negative_best_query_coverage,
        "negative_best_evalue": record.negative_best_evalue,
        "negative_best_source": record.negative_best_source,
        "negative_hit_strength": record.negative_hit_strength or "",
        "negative_strong_hit_count": record.negative_strong_hit_count,
        "negative_medium_hit_count": record.negative_medium_hit_count,
        "negative_weak_hit_count": record.negative_weak_hit_count,
        "negative_exclusion_reason": record.negative_exclusion_reason,
        "motifs": "; ".join(record.motifs),
        "uniprot_accession": record.uniprot_accession,
        "alphafold_url": record.alphafold_url,
        "notes": "; ".join(record.notes),
        "positive_source_count": record.positive_source_count,
        "positive_sources_hit": "; ".join(record.positive_sources_hit),
        "positive_sources_missing": "; ".join(record.positive_sources_missing),
    }


def _score_components(record: ProteinRecord) -> str:
    """Return score components as readable name=value pairs."""
    if record.score is None:
        return ""

    return "; ".join(
        f"{name}={value}" for name, value in record.score.components.items()
    )


def _unique_domain_count(record: ProteinRecord) -> int:
    """Return the number of unique domain accessions in first-seen order."""
    return len(
        {
            str(domain.accession)
            for domain in record.domains
            if str(domain.accession)
        }
    )


def _unique_domain_names(record: ProteinRecord) -> str:
    """Return unique readable domain names while skipping internal numeric ids."""
    return _unique_join(
        domain.name
        for domain in record.domains
        if not _looks_like_internal_domain_name(domain.name)
    )


def _looks_like_internal_domain_name(value: object) -> bool:
    """Return True for numeric/internal-looking domain names."""
    return bool(re.fullmatch(r"\d{6,}", str(value).strip()))


def _best_hit(hits: list[BlastHit]) -> BlastHit | None:
    """Return the best BLAST hit by highest bitscore, then lowest e-value."""
    if not hits:
        return None

    return max(hits, key=lambda hit: (hit.bitscore, -hit.evalue))


def _join(values: Iterable[object]) -> str:
    """Join non-empty string values with a semicolon separator."""
    return "; ".join(str(value) for value in values if str(value))


def _unique_join(values: Iterable[object]) -> str:
    """Join unique non-empty string values while preserving input order."""
    seen: set[str] = set()
    unique_values: list[str] = []

    for value in values:
        text = str(value)
        if not text or text in seen:
            continue

        seen.add(text)
        unique_values.append(text)

    return "; ".join(unique_values)


def _blast_status(record: ProteinRecord) -> str:
    """Return a compact BLAST hit status for Excel output."""
    has_positive = bool(record.positive_hits)
    has_negative = bool(record.negative_hits)

    if has_positive and has_negative:
        return "positive_and_negative"

    if has_positive:
        return "positive_only"

    if has_negative:
        return "negative_only"

    return "no_hits"


__all__: tuple[str, ...] = (
    "CANDIDATE_OVERVIEW_COLUMNS",
    "CATEGORY_EVIDENCE_SHEETS",
    "EXCEL_COLUMNS",
    "FINAL_SCORE_COLUMNS",
    "INDEX_ROWS_V2",
    "INTERACTION_SCORE_EXPLANATIONS",
    "INTERACTION_SCORE_NOTES",
    "SCORE_BREAKDOWN_COLUMNS",
    "records_to_dataframe",
    "write_classification_workbook",
    "write_records_to_excel",
)
